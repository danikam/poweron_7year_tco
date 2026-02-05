"""
TCO Calculator V2 - Parameter Loader
Extracts all parameters from Excel for faithful replication
"""

import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import Optional


@dataclass
class TCOParameters:
    """All parameters needed for TCO calculations - extracted from Excel"""
    
    # Vehicle specifications
    ice_msrp: float
    ev_msrp: float
    ice_efficiency_l_per_km: float
    ev_efficiency_kwh_per_km: float
    annual_km: float
    
    # Fuel/Energy costs
    gasoline_price_per_l: float
    electricity_price_per_kwh: float
    fuel_price_escalation: float
    electricity_price_escalation: float
    
    # Maintenance
    ice_maintenance_per_km: float
    ev_maintenance_per_km: float
    maintenance_escalation: float
    charger_maintenance_rate: float  # % of charger cost
    
    # Insurance
    ice_insurance_rate: float  # % of MSRP
    ev_insurance_rate: float   # % of MSRP
    
    # Depreciation
    ice_depreciation_rate: float
    ev_depreciation_rate: float
    charger_depreciation_rate: float  # Class 43.1 or 43.2
    
    # Charger infrastructure
    charger_cost: float
    charger_cost_per_vehicle: float  # charger_cost / num_vehicles
    charger_depreciation_rate: float  # B27 (0.3) or B28 (0.5) based on charger type
    
    # Incentives & Credits
    ev_federal_rebate: float  # iMHZEV
    charger_incentives: float
    cfr_credit_annual: float
    lcfs_credit_annual: float
    lcfs_decline_rate: float
    credit_end_year: int  # Last year for CFR/LCFS
    
    # Carbon tax
    carbon_tax_per_l: float
    
    # Financial
    analysis_years: int  # Usually 7
    discount_rate: float
    
    # Financing (Scenarios 2 & 3)
    ice_down_payment_pct: Optional[float] = None
    ice_interest_rate: Optional[float] = None
    ice_loan_term_years: Optional[int] = None
    ev_down_payment_pct: Optional[float] = None
    ev_interest_rate: Optional[float] = None
    ev_loan_term_years: Optional[int] = None
    
    # Charging-as-a-Service (Scenario 3)
    caas_markup: Optional[float] = None  # Default 1.8 (80% markup)
    caas_term_years: Optional[int] = None  # Default 7


def _get_charger_depreciation_rate(wb: openpyxl.Workbook) -> float:
    """Determine charger depreciation rate (B27=0.3 or B28=0.5) based on charger type
    
    The formula checks: IF('TCO Dashboard'!C20 = 'Charger database'!A6, B28, B27)
    This checks if the selected charger equals row 6 (L3 high power 150kW)
    If yes, use B28 (0.5), else use B27 (0.3)
    """
    calc_sheet = wb["Cost analysis - purchase"]
    b27 = calc_sheet.cell(row=27, column=2).value or 0.3
    b28 = calc_sheet.cell(row=28, column=2).value or 0.5
    
    try:
        tco_sheet = wb["TCO Dashboard"]
        charger_sheet = wb["Charger database"]
        
        tco_c20 = tco_sheet.cell(row=20, column=3).value
        charger_a6 = charger_sheet.cell(row=6, column=1).value
        
        # Return B28 if match, else B27
        return b28 if (tco_c20 == charger_a6) else b27
    except:
        # Default to B27 (Class 43.1 = 0.3)
        return b27


def load_parameters_from_excel(excel_path: Path, scenario: int = 1) -> TCOParameters:
    """
    Load all parameters from Excel file
    
    Args:
        excel_path: Path to 7_year_TCO_Canada.xlsx
        scenario: 1 (Purchase), 2 (Finance), or 3 (CaaS)
    
    Returns:
        TCOParameters with all values extracted from Excel
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    calc_sheet = wb["Cost analysis - purchase"]
    
    def get_param(ref: str) -> float:
        """Get parameter value from B column"""
        row = int(ref[1:])
        value = calc_sheet.cell(row=row, column=2).value
        if value is None:
            return 0.0
        return float(value)
    
    # Extract all parameters
    params = TCOParameters(
        # Vehicle specs
        ice_msrp=get_param("B6"),
        ev_msrp=get_param("B7"),
        ice_efficiency_l_per_km=get_param("B20"),
        ev_efficiency_kwh_per_km=get_param("B19"),
        annual_km=get_param("B21"),
        
        # Energy costs
        gasoline_price_per_l=get_param("B9"),
        electricity_price_per_kwh=get_param("B10"),
        fuel_price_escalation=get_param("B11"),
        electricity_price_escalation=get_param("B12"),
        
        # Maintenance
        ice_maintenance_per_km=get_param("B14"),
        ev_maintenance_per_km=get_param("B15"),
        maintenance_escalation=get_param("B16"),
        charger_maintenance_rate=get_param("B17"),
        
        # Insurance
        ice_insurance_rate=get_param("B55"),
        ev_insurance_rate=get_param("B54"),
        
        # Depreciation
        ice_depreciation_rate=get_param("B25"),
        ev_depreciation_rate=get_param("B26"),
        
        # Charger - determine correct depreciation rate based on charger type
        charger_cost=get_param("B37"),
        charger_cost_per_vehicle=get_param("B38"),
        charger_depreciation_rate=_get_charger_depreciation_rate(wb),
        
        # Incentives
        ev_federal_rebate=get_param("B50"),
        charger_incentives=get_param("B52"),
        cfr_credit_annual=get_param("B41"),
        lcfs_credit_annual=get_param("B43"),
        lcfs_decline_rate=get_param("B47"),
        credit_end_year=int(get_param("B42")),
        
        # Carbon tax
        carbon_tax_per_l=get_param("B40"),
        
        # Financial
        analysis_years=7,  # Standard 7-year analysis
        discount_rate=get_param("B2"),  # Discount rate for NPV calculations
    )
    
    # Load financing parameters for Scenarios 2 & 3
    if scenario in [2, 3]:
        params.ice_down_payment_pct = get_param("B30")
        params.ice_interest_rate = get_param("B31")
        params.ice_loan_term_years = int(get_param("B32"))
        params.ev_down_payment_pct = get_param("B33")
        params.ev_interest_rate = get_param("B34")
        params.ev_loan_term_years = int(get_param("B35"))
    
    # Load CaaS parameters for Scenario 3
    if scenario == 3:
        params.caas_markup = 1.8  # 80% markup from formula
        params.caas_term_years = 7  # Standard term
    
    wb.close()
    return params


def validate_parameters(params: TCOParameters, scenario: int) -> list[str]:
    """
    Validate that all required parameters are present and reasonable
    
    Returns:
        List of validation errors (empty if all OK)
    """
    errors = []
    
    # Check positive values
    if params.ice_msrp <= 0:
        errors.append("ICE MSRP must be positive")
    if params.ev_msrp <= 0:
        errors.append("EV MSRP must be positive")
    if params.annual_km <= 0:
        errors.append("Annual KM must be positive")
    
    # Check reasonable ranges
    if params.ice_efficiency_l_per_km > 1.0:
        errors.append(f"ICE efficiency unusually high: {params.ice_efficiency_l_per_km} L/km")
    if params.ev_efficiency_kwh_per_km > 2.0:
        errors.append(f"EV efficiency unusually high: {params.ev_efficiency_kwh_per_km} kWh/km")
    
    # Check financing parameters for Scenarios 2 & 3
    if scenario in [2, 3]:
        if params.ice_down_payment_pct is None or params.ice_down_payment_pct < 0:
            errors.append("ICE down payment % required for financing scenario")
        if params.ice_interest_rate is None or params.ice_interest_rate < 0:
            errors.append("ICE interest rate required for financing scenario")
        if params.ice_loan_term_years is None or params.ice_loan_term_years <= 0:
            errors.append("ICE loan term required for financing scenario")
    
    return errors


if __name__ == "__main__":
    # Test parameter loading
    excel_path = Path("/Users/danikae/Git/tco_model/excel_tool/7_year_TCO_Canada.xlsx")
    
    print("Loading parameters for Scenario 1...")
    params = load_parameters_from_excel(excel_path, scenario=1)
    
    print(f"\n✅ Parameters loaded:")
    print(f"  ICE MSRP: ${params.ice_msrp:,.2f}")
    print(f"  EV MSRP: ${params.ev_msrp:,.2f}")
    print(f"  Annual KM: {params.annual_km:,.0f}")
    print(f"  Gasoline: ${params.gasoline_price_per_l:.3f}/L")
    print(f"  Electricity: ${params.electricity_price_per_kwh:.3f}/kWh")
    print(f"  Charger cost: ${params.charger_cost:,.2f}")
    print(f"  EV rebate: ${params.ev_federal_rebate:,.2f}")
    print(f"  Discount rate: {params.discount_rate*100:.1f}%")
    
    errors = validate_parameters(params, scenario=1)
    if errors:
        print(f"\n❌ Validation errors:")
        for error in errors:
            print(f"  - {error}")
    else:
        print(f"\n✅ All parameters validated successfully")
